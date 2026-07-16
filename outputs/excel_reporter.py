# -*- coding: utf-8 -*-
"""
Excel 输出模块
生成每日内容排期表
"""
import os
from datetime import datetime, timedelta
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from utils.file_manager import get_file_manager
from utils.logger import logger


class ExcelReporter:
    """生成 Excel 内容排期表"""
    
    def __init__(self):
        self.fm = get_file_manager()
    
    def create_daily_schedule(self, 
                              topics: List[Dict], 
                              contents: List[Dict],
                              date_str: str) -> str:
        """
        生成每日内容排期表
        包含：热点、选题、平台、内容类型、负责人、状态、发布时间
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "每日内容排期"
        
        # 样式定义
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(bold=True, size=14)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 标题行
        ws.merge_cells("A1:H1")
        ws["A1"] = f"每日热点内容营销排期表 - {date_str}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # 表头
        headers = ["序号", "热点主题", "匹配度", "适用平台", "内容形式", 
                   "执行人", "计划发布时间", "状态"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # 数据行
        row = 4
        for idx, topic in enumerate(topics, 1):
            # 每个选题在多个平台排期
            platforms = ["小红书", "抖音", "站内"]
            for platform in platforms:
                ws.cell(row=row, column=1, value=idx).border = thin_border
                ws.cell(row=row, column=2, value=topic["hot_topic"]["title"]).border = thin_border
                
                # 匹配度颜色
                score = topic["match_score"]
                score_cell = ws.cell(row=row, column=3, value=f"{score}分")
                score_cell.border = thin_border
                if score >= 70:
                    score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif score >= 40:
                    score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                ws.cell(row=row, column=4, value=platform).border = thin_border
                
                # 根据平台推荐内容形式
                if platform == "小红书":
                    fmt = "图文种草"
                elif platform == "抖音":
                    fmt = "短视频"
                else:
                    fmt = "站内活动/Banner"
                ws.cell(row=row, column=5, value=fmt).border = thin_border
                
                ws.cell(row=row, column=6, value="待分配").border = thin_border
                
                # 计划发布时间（根据平台错峰）
                if platform == "小红书":
                    pub_time = f"{date_str} 12:00"
                elif platform == "抖音":
                    pub_time = f"{date_str} 18:00"
                else:
                    pub_time = f"{date_str} 10:00"
                ws.cell(row=row, column=7, value=pub_time).border = thin_border
                
                ws.cell(row=row, column=8, value="待执行").border = thin_border
                
                row += 1
        
        # 列宽
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 10
        
        # 保存
        filepath = self.fm.get_filename(date_str, "内容排期表", ".xlsx")
        wb.save(filepath)
        logger.info(f"[Excel] 排期表已生成: {filepath}")
        return filepath


# 单例
_excel_reporter = None

def get_excel_reporter() -> ExcelReporter:
    global _excel_reporter
    if _excel_reporter is None:
        _excel_reporter = ExcelReporter()
    return _excel_reporter
